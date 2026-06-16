import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Khmer Font — cached locally so PDF generation works offline after first run
# ---------------------------------------------------------------------------
_FONTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
_FONT_PATH = os.path.join(_FONTS_DIR, "NotoSansKhmer-Regular.ttf")

_FONT_URLS = [
    # Noto Sans Khmer — best Khmer coverage from Google Fonts
    "https://fonts.gstatic.com/s/notosanskhmer/v18/ijw3s5roRME5LLRxjsRb-gssOmMaue_tpFE.ttf",
    # Fallback: Nokora
    "https://fonts.gstatic.com/s/nokora/v32/hESw6XVnNCxEvkbMpheEZo_H_w.ttf",
]

async def ensure_khmer_font_async() -> str:
    """Check if Khmer font is cached locally. Returns path to font or empty string."""
    if os.path.exists(_FONT_PATH):
        return _FONT_PATH
    return ""


def _ensure_khmer_font() -> str:
    """Check if Khmer font is cached locally. Returns path to font or empty string."""
    if os.path.exists(_FONT_PATH):
        return _FONT_PATH
    return ""


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

from database import parse_note_details

def get_report_period_string(reports_data: list) -> str:
    if not reports_data:
        return "N/A"
    from database import extract_report_day
    start_day = extract_report_day(reports_data[0]['report'].header)
    end_day = extract_report_day(reports_data[-1]['report'].header)
    return start_day if start_day == end_day else f"{start_day}_to_{end_day}"


def aggregate_summary_data(reports_data: list) -> dict:
    summary = {}
    for data in reports_data:
        for record in data['records']:
            name = record.employee_name
            gender_val = getattr(record, 'gender', '') or ''
            borrow_val, deduct_val, _ = parse_note_details(record.note)
            ot_hours = max(0.0, record.hours - 8.0)
            if name not in summary:
                summary[name] = {
                    'hours': 0.0,
                    'salary': 0.0,
                    'ot_hours': 0.0,
                    'borrow': 0.0,
                    'deduction': 0.0,
                    'rate': record.daily_rate,
                    'gender': gender_val
                }
            summary[name]['hours'] += record.hours
            summary[name]['salary'] += record.salary
            summary[name]['ot_hours'] += ot_hours
            summary[name]['borrow'] += borrow_val
            summary[name]['deduction'] += deduct_val
            if record.daily_rate > 0:
                summary[name]['rate'] = record.daily_rate
            if gender_val:
                summary[name]['gender'] = gender_val
    return summary


# ---------------------------------------------------------------------------
# Excel Report
# ---------------------------------------------------------------------------

def generate_excel_report(reports_data: list, output_path: str, exchange_rate: float = 4000.0):
    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # ---- 1. Summary Sheet ----
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    ws_summary.freeze_panes = 'E5'

    ws_summary.merge_cells("A1:K1")
    ws_summary["A1"] = "សរុបប្រាក់ឈ្នួល និងវត្តមាន"
    ws_summary["A1"].font = Font(name="Times New Roman", size=16, bold=True, color="1F497D")
    ws_summary["A1"].alignment = Alignment(horizontal="center")

    ws_summary["A2"] = f"រយៈពេល: {get_report_period_string(reports_data)}"
    ws_summary["A2"].font = Font(name="Times New Roman", size=11, italic=True)

    headers_summary = [
        "ឈ្មោះបុគ្គលិក",
        "ភេទ",
        "តម្លៃថ្ងៃ (KHR)",
        "តម្លៃម៉ោង (KHR)",
        "សរុបថ្ងៃធ្វើការ",
        "សរុបថែមម៉ោង",
        "លុយសរុប (KHR)",
        "លុយបានខ្ចី សរុប (KHR)",
        "ប្រាក់កាត់ (KHR)",
        "លុយត្រូវបើក (KHR)",
        "លុយត្រូវបើក (USD)"
    ]
    for col_idx, header in enumerate(headers_summary, 1):
        cell = ws_summary.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    summary_data = aggregate_summary_data(reports_data)
    row_idx = 5
    for name, stats in sorted(summary_data.items()):
        base_days = (stats['hours'] - stats['ot_hours']) / 8.0
        
        ws_summary.cell(row=row_idx, column=1, value=name)
        ws_summary.cell(row=row_idx, column=2, value=stats.get('gender', '')) # ភេទ
        ws_summary.cell(row=row_idx, column=3, value=stats['rate']).number_format = '#,##0" ៛"'
        ws_summary.cell(row=row_idx, column=4, value=f"=C{row_idx}/8").number_format = '#,##0" ៛"'
        ws_summary.cell(row=row_idx, column=5, value=base_days).number_format = '#,##0.0'
        ws_summary.cell(row=row_idx, column=6, value=stats['ot_hours']).number_format = '#,##0.0'
        
        # Formulas for salaries
        ws_summary.cell(row=row_idx, column=7, value=f"=(E{row_idx}*C{row_idx})+(F{row_idx}*D{row_idx})").number_format = '#,##0" ៛"'
        ws_summary.cell(row=row_idx, column=8, value=stats['borrow']).number_format = '#,##0" ៛"'
        ws_summary.cell(row=row_idx, column=9, value=stats['deduction']).number_format = '#,##0" ៛"'
        ws_summary.cell(row=row_idx, column=10, value=f"=G{row_idx}-H{row_idx}-I{row_idx}").number_format = '#,##0" ៛"'
        ws_summary.cell(row=row_idx, column=11, value=f"=J{row_idx}/{exchange_rate}").number_format = '$#,##0.00'
        
        for c in range(1, 12):
            cell = ws_summary.cell(row=row_idx, column=c)
            cell.border = thin_border
            cell.font = Font(name="Times New Roman")
            if c in [1, 2]:
                cell.alignment = Alignment(horizontal="center" if c == 2 else "left")
            else:
                cell.alignment = Alignment(horizontal="right")
        row_idx += 1

    ws_summary.cell(row=row_idx, column=1, value="សរុប").font = Font(name="Times New Roman", bold=True, size=11)
    
    ws_summary.cell(row=row_idx, column=5, value=f"=SUM(E5:E{row_idx-1})").font = Font(name="Times New Roman", bold=True)
    ws_summary.cell(row=row_idx, column=5).number_format = '#,##0.0'
    
    ws_summary.cell(row=row_idx, column=6, value=f"=SUM(F5:F{row_idx-1})").font = Font(name="Times New Roman", bold=True)
    ws_summary.cell(row=row_idx, column=6).number_format = '#,##0.0'
    
    ws_summary.cell(row=row_idx, column=7, value=f"=SUM(G5:G{row_idx-1})").font = Font(name="Times New Roman", bold=True)
    ws_summary.cell(row=row_idx, column=7).number_format = '#,##0" ៛"'
    
    ws_summary.cell(row=row_idx, column=8, value=f"=SUM(H5:H{row_idx-1})").font = Font(name="Times New Roman", bold=True)
    ws_summary.cell(row=row_idx, column=8).number_format = '#,##0" ៛"'
    
    ws_summary.cell(row=row_idx, column=9, value=f"=SUM(I5:I{row_idx-1})").font = Font(name="Times New Roman", bold=True)
    ws_summary.cell(row=row_idx, column=9).number_format = '#,##0" ៛"'
    
    ws_summary.cell(row=row_idx, column=10, value=f"=SUM(J5:J{row_idx-1})").font = Font(name="Times New Roman", bold=True)
    ws_summary.cell(row=row_idx, column=10).number_format = '#,##0" ៛"'
    
    ws_summary.cell(row=row_idx, column=11, value=f"=SUM(K5:K{row_idx-1})").font = Font(name="Times New Roman", bold=True)
    ws_summary.cell(row=row_idx, column=11).number_format = '$#,##0.00'
    
    double_bottom = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    for c in range(1, 12):
        cell = ws_summary.cell(row=row_idx, column=c)
        cell.border = double_bottom
        if c > 1:
            cell.alignment = Alignment(horizontal="right")

    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ---- 2. Details Sheet ----
    ws_details = wb.create_sheet(title="Details")
    ws_details.views.sheetView[0].showGridLines = True
    ws_details.freeze_panes = 'F5'

    from database import extract_report_day, parse_date, get_employee_rate

    # Determine unique date columns from reports_data
    date_columns = []
    for item in reports_data:
        rep = item['report']
        day_str = extract_report_day(rep.header)
        d = parse_date(day_str)
        col_header = str(d.day) if d else day_str
        date_columns.append({
            'report_id': rep.id,
            'header': col_header,
            'full_date': day_str
        })

    K = len(date_columns)
    col_days = 6 + 4*K
    col_ot = col_days + 1
    col_gross = col_days + 2
    col_borrow = col_days + 3
    col_deduct = col_days + 4
    col_net = col_days + 5

    last_col_letter = get_column_letter(col_net)

    ws_details.merge_cells(f"A1:{last_col_letter}1")
    ws_details["A1"] = "ព័ត៌មានលម្អិតប្រចាំថ្ងៃ"
    ws_details["A1"].font = Font(name="Times New Roman", size=16, bold=True, color="1F497D")
    ws_details["A1"].alignment = Alignment(horizontal="center")

    ws_details["A2"] = f"រយៈពេល: {get_report_period_string(reports_data)}"
    ws_details["A2"].font = Font(name="Times New Roman", size=11, italic=True)

    header_font = Font(name="Times New Roman", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Style all header cells first
    for r in [3, 4]:
        for c in range(1, col_net + 1):
            cell = ws_details.cell(row=r, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # Merge vertical headers on left
    ws_details.merge_cells("A3:A4")
    ws_details.cell(row=3, column=1, value="លរ")
    ws_details.merge_cells("B3:B4")
    ws_details.cell(row=3, column=2, value="ឈ្មោះ")
    ws_details.merge_cells("C3:C4")
    ws_details.cell(row=3, column=3, value="ភេទ")
    ws_details.merge_cells("D3:D4")
    ws_details.cell(row=3, column=4, value="តម្លៃ(៛)/ថ្ងៃ")
    ws_details.merge_cells("E3:E4")
    ws_details.cell(row=3, column=5, value="តម្លៃ(៛)/ម៉ោង")

    # Merge dates and set subheaders
    c_idx = 6
    for date_col in date_columns:
        start_letter = get_column_letter(c_idx)
        end_letter = get_column_letter(c_idx + 3)
        ws_details.merge_cells(f"{start_letter}3:{end_letter}3")
        ws_details.cell(row=3, column=c_idx, value=date_col['header'])
        
        ws_details.cell(row=4, column=c_idx, value="Day")
        ws_details.cell(row=4, column=c_idx + 1, value="OT(h)")
        ws_details.cell(row=4, column=c_idx + 2, value="Borrow")
        ws_details.cell(row=4, column=c_idx + 3, value="Project")
        c_idx += 4

    # Merge summary columns on right
    summary_headers = [
        "សរុបថ្ងៃធ្វើការ",
        "សរុបថែមម៉ោង",
        "លុយសរុប",
        "លុយបានខ្ចី សរុប",
        "ប្រាក់កាត់",
        "លុយត្រូវបើក"
    ]
    for idx, h in enumerate(summary_headers):
        curr_c = col_days + idx
        col_letter = get_column_letter(curr_c)
        ws_details.merge_cells(f"{col_letter}3:{col_letter}4")
        ws_details.cell(row=3, column=curr_c, value=h)

    # Collect unique employee names
    employees = sorted(list(set(
        record.employee_name
        for item in reports_data
        for record in item['records']
    )))

    # Map hours, notes and gender to (employee_name, report_id) or employee name
    hours_map = {}
    notes_map = {}
    employee_gender = {}
    for item in reports_data:
        rep_id = item['report'].id
        for record in item['records']:
            key = (record.employee_name, rep_id)
            hours_map[key] = record.hours
            notes_map[key] = record.note
            g_val = getattr(record, 'gender', '') or ''
            if g_val:
                employee_gender[record.employee_name] = g_val

    det_row = 5
    for idx, name in enumerate(employees, 1):
        ws_details.cell(row=det_row, column=1, value=idx)
        ws_details.cell(row=det_row, column=2, value=name)
        ws_details.cell(row=det_row, column=3, value=employee_gender.get(name, "")) # ភេទ
        
        rate = get_employee_rate(name) or 0.0
        ws_details.cell(row=det_row, column=4, value=rate).number_format = '#,##0" ៛"'
        ws_details.cell(row=det_row, column=5, value=f"=D{det_row}/8").number_format = '#,##0" ៛"'
        
        c_idx = 6
        total_deductions_val = 0.0
        for date_col in date_columns:
            rep_id = date_col['report_id']
            key = (name, rep_id)
            hours = hours_map.get(key, None)
            note = notes_map.get(key, None)
            
            borrow_val = 0.0
            deduct_val = 0.0
            cleaned_note = ""
            if note:
                borrow_val, deduct_val, cleaned_note = parse_note_details(note)
                total_deductions_val += deduct_val
                
            if hours is not None:
                day_val = min(8.0, hours) / 8.0
                ot_val = max(0.0, hours - 8.0)
                ws_details.cell(row=det_row, column=c_idx, value=day_val).number_format = '#,##0.0'
                ws_details.cell(row=det_row, column=c_idx + 1, value=ot_val).number_format = '#,##0.0'
                ws_details.cell(row=det_row, column=c_idx + 2, value=borrow_val if borrow_val > 0 else "").number_format = '#,##0" ៛"'
                ws_details.cell(row=det_row, column=c_idx + 3, value=cleaned_note)
            else:
                ws_details.cell(row=det_row, column=c_idx, value="")
                ws_details.cell(row=det_row, column=c_idx + 1, value="")
                ws_details.cell(row=det_row, column=c_idx + 2, value="")
                ws_details.cell(row=det_row, column=c_idx + 3, value="")
            c_idx += 4

        # Formulas for summary columns on details sheet
        L_days = get_column_letter(col_days)
        L_ot = get_column_letter(col_ot)
        L_gross = get_column_letter(col_gross)
        L_borrow = get_column_letter(col_borrow)
        L_deduct = get_column_letter(col_deduct)
        L_net = get_column_letter(col_net)

        days_sum_parts = [f"{get_column_letter(c)}{det_row}" for c in range(6, 6 + 4*K, 4)]
        days_formula = f"={'+'.join(days_sum_parts)}" if days_sum_parts else "=0"
        
        ot_sum_parts = [f"{get_column_letter(c + 1)}{det_row}" for c in range(6, 6 + 4*K, 4)]
        ot_formula = f"={'+'.join(ot_sum_parts)}" if ot_sum_parts else "=0"
        
        gross_formula = f"=({L_days}{det_row}*D{det_row})+({L_ot}{det_row}*E{det_row})"
        
        borrow_sum_parts = [f"{get_column_letter(c + 2)}{det_row}" for c in range(6, 6 + 4*K, 4)]
        borrow_formula = f"={'+'.join(borrow_sum_parts)}" if borrow_sum_parts else "=0"
        
        net_formula = f"={L_gross}{det_row}-{L_borrow}{det_row}-{L_deduct}{det_row}"

        ws_details.cell(row=det_row, column=col_days, value=days_formula).number_format = '#,##0.0'
        ws_details.cell(row=det_row, column=col_ot, value=ot_formula).number_format = '#,##0.0'
        ws_details.cell(row=det_row, column=col_gross, value=gross_formula).number_format = '#,##0" ៛"'
        ws_details.cell(row=det_row, column=col_borrow, value=borrow_formula).number_format = '#,##0" ៛"'
        ws_details.cell(row=det_row, column=col_deduct, value=total_deductions_val).number_format = '#,##0" ៛"'
        ws_details.cell(row=det_row, column=col_net, value=net_formula).number_format = '#,##0" ៛"'

        # Styling
        for c in range(1, col_net + 1):
            cell = ws_details.cell(row=det_row, column=c)
            cell.border = thin_border
            cell.font = Font(name="Times New Roman")
            if c in [1, 3]:
                cell.alignment = Alignment(horizontal="center")
            elif c == 2 or (c >= 6 and (c - 6) % 4 == 3):
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
        det_row += 1

    # Total Row
    ws_details.cell(row=det_row, column=1, value="សរុប")
    
    for i in range(K):
        base_c = 6 + 4*i
        # Day sum
        col_let = get_column_letter(base_c)
        ws_details.cell(row=det_row, column=base_c, value=f"=SUM({col_let}5:{col_let}{det_row-1})").number_format = '#,##0.0'
        
        # OT sum
        col_let = get_column_letter(base_c + 1)
        ws_details.cell(row=det_row, column=base_c + 1, value=f"=SUM({col_let}5:{col_let}{det_row-1})").number_format = '#,##0.0'
        
        # Borrow sum
        col_let = get_column_letter(base_c + 2)
        ws_details.cell(row=det_row, column=base_c + 2, value=f"=SUM({col_let}5:{col_let}{det_row-1})").number_format = '#,##0" ៛"'
        
    for c in [col_days, col_ot, col_gross, col_borrow, col_deduct, col_net]:
        col_let = get_column_letter(c)
        ws_details.cell(row=det_row, column=c, value=f"=SUM({col_let}5:{col_let}{det_row-1})")
        if c in [col_days, col_ot]:
            ws_details.cell(row=det_row, column=c).number_format = '#,##0.0'
        else:
            ws_details.cell(row=det_row, column=c).number_format = '#,##0" ៛"'

    double_bottom = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    for c in range(1, col_net + 1):
        cell = ws_details.cell(row=det_row, column=c)
        cell.border = double_bottom
        cell.font = Font(name="Times New Roman", bold=True)
        if c >= 4:
            cell.alignment = Alignment(horizontal="right")
        elif c in [1, 3]:
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="left")

    # Column Widths
    for col in ws_details.columns:
        col_idx_val = col[0].column
        col_letter = get_column_letter(col_idx_val)
        if col_idx_val == 1:
            ws_details.column_dimensions[col_letter].width = 6
        elif col_idx_val == 2:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws_details.column_dimensions[col_letter].width = max(max_len + 3, 16)
        elif col_idx_val == 3:
            ws_details.column_dimensions[col_letter].width = 6
        elif col_idx_val in [4, 5]:
            ws_details.column_dimensions[col_letter].width = 15
        elif col_idx_val >= 6 and col_idx_val < col_days:
            sub_col_type = (col_idx_val - 6) % 4
            if sub_col_type == 3:
                ws_details.column_dimensions[col_letter].width = 15
            else:
                ws_details.column_dimensions[col_letter].width = 10
        else:
            ws_details.column_dimensions[col_letter].width = 16

    wb.save(output_path)


# ---------------------------------------------------------------------------
# PDF Report (WeasyPrint — proper Khmer shaping via Pango)
# ---------------------------------------------------------------------------

def _build_report_html(reports_data: list, period_str: str, font_path: str, exchange_rate: float = 4000.0) -> str:
    """Build a complete HTML document for the report."""
    summary_data = aggregate_summary_data(reports_data)
    
    total_days_sum = 0.0
    total_ot_sum = 0.0
    total_gross_sum = 0.0
    total_borrow_sum = 0.0
    total_deduct_sum = 0.0

    # Summary table rows
    summary_rows = ""
    for i, (name, stats) in enumerate(sorted(summary_data.items()), 1):
        row_class = "even" if i % 2 == 0 else ""
        base_days = (stats['hours'] - stats['ot_hours']) / 8.0
        ot_hours = stats['ot_hours']
        
        hourly_rate = stats['rate'] / 8.0
        gross_salary = base_days * stats['rate'] + ot_hours * hourly_rate
        net_salary = gross_salary - stats['borrow'] - stats['deduction']
        net_salary_usd = net_salary / exchange_rate
        
        total_days_sum += base_days
        total_ot_sum += ot_hours
        total_gross_sum += gross_salary
        total_borrow_sum += stats['borrow']
        total_deduct_sum += stats['deduction']

        summary_rows += f"""
            <tr class="{row_class}">
                <td>{i}</td>
                <td class="name">{name}</td>
                <td class="center">{stats.get('gender', '')}</td>
                <td class="num">{stats['rate']:,.0f} ៛</td>
                <td class="num">{hourly_rate:,.0f} ៛</td>
                <td class="num">{base_days:.1f} ថ្ងៃ</td>
                <td class="num">{ot_hours:.1f} h</td>
                <td class="num">{int(round(gross_salary)):,} ៛</td>
                <td class="num">{int(round(stats['borrow'])):,} ៛</td>
                <td class="num">{int(round(stats['deduction'])):,} ៛</td>
                <td class="num">${net_salary_usd:.2f}</td>
                <td class="num">{int(round(net_salary)):,} ៛</td>
            </tr>"""

    total_net_khr_sum = total_gross_sum - total_borrow_sum - total_deduct_sum
    total_net_usd_sum = total_net_khr_sum / exchange_rate

    # Font URI
    font_css = ""
    if font_path and os.path.exists(font_path):
        import base64
        with open(font_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("ascii")
        font_css = f"""@font-face {{
            font-family: 'KhmerFont';
            src: url('data:font/truetype;base64,{b64}');
        }}"""

    # Daily breakdown sections
    daily_sections = ""
    for day_data in reports_data:
        report = day_data['report']
        records = day_data['records']
        
        day_total_days = 0.0
        day_total_ot = 0.0
        day_total_gross = 0.0
        day_total_borrow = 0.0
        day_total_deduct = 0.0
        day_total_net = 0.0

        rows = ""
        for i, rec in enumerate(records, 1):
            row_class = "even" if i % 2 == 0 else ""
            borrow_val, deduct_val, cleaned_note = parse_note_details(rec.note)
            
            day_val = min(8.0, rec.hours) / 8.0
            ot_val = max(0.0, rec.hours - 8.0)
            
            h_rate = rec.daily_rate / 8.0
            gross_val = day_val * rec.daily_rate + ot_val * h_rate
            net_val = gross_val - borrow_val - deduct_val
            
            day_total_days += day_val
            day_total_ot += ot_val
            day_total_gross += gross_val
            day_total_borrow += borrow_val
            day_total_deduct += deduct_val
            day_total_net += net_val

            rows += f"""
                <tr class="{row_class}">
                    <td class="center">{i}</td>
                    <td class="name">{rec.employee_name}</td>
                    <td class="center">{getattr(rec, 'gender', '') or ''}</td>
                    <td class="num">{day_val:.1f}</td>
                    <td class="num">{ot_val:.1f}</td>
                    <td class="num">{rec.daily_rate:,.0f} ៛</td>
                    <td class="num">{int(round(gross_val)):,} ៛</td>
                    <td class="num">{int(round(borrow_val)):,} ៛</td>
                    <td class="num">{int(round(deduct_val)):,} ៛</td>
                    <td class="num">{int(round(net_val)):,} ៛</td>
                    <td>{cleaned_note}</td>
                </tr>"""

        daily_sections += f"""
            <div class="day-block">
                <h3>{report.header}</h3>
                <table>
                    <thead>
                        <tr>
                            <th style="width:30px">ល.រ</th>
                            <th>ឈ្មោះបុគ្គលិក</th>
                            <th>ភេទ</th>
                            <th class="num">ថ្ងៃធ្វើការ</th>
                            <th class="num">ថែមម៉ោង</th>
                            <th class="num">តម្លៃថ្ងៃ</th>
                            <th class="num">លុយសរុប</th>
                            <th class="num">ខ្ចី</th>
                            <th class="num">កាត់</th>
                            <th class="num">ត្រូវបើក</th>
                            <th>សម្គាល់</th>
                        </tr>
                    </thead>
                    <tbody>{rows}</tbody>
                    <tfoot>
                        <tr class="subtotal">
                            <td colspan="3">សរុបថ្ងៃ</td>
                            <td class="num">{day_total_days:.1f}</td>
                            <td class="num">{day_total_ot:.1f}</td>
                            <td></td>
                            <td class="num">{int(round(day_total_gross)):,} ៛</td>
                            <td class="num">{int(round(day_total_borrow)):,} ៛</td>
                            <td class="num">{int(round(day_total_deduct)):,} ៛</td>
                            <td class="num">{int(round(day_total_net)):,} ៛</td>
                            <td></td>
                        </tr>
                    </tfoot>
                </table>
            </div>"""

    display_period = period_str.replace("_to_", " ដល់ ")

    return f"""<!DOCTYPE html>
<html lang="km">
<head>
<meta charset="utf-8">
<title>របាយការណ៍វត្តមាន — {display_period}</title>
<style>
{font_css}

@page {{
    size: A4 portrait;
    margin: 15mm;
}}

@page landscapePage {{
    size: A4 landscape;
    margin: 15mm;
}}

.first-page {{
    page: landscapePage;
}}

* {{ box-sizing: border-box; margin: 0; padding: 0; }}

body {{
    font-family: 'KhmerFont', 'Noto Sans Khmer', 'Khmer MN', 'Khmer Sangam MN', sans-serif;
    font-size: 8pt;
    color: #222;
    padding: 15px 20px;
    line-height: 1.4;
}}

/* ---- Header ---- */
.report-title {{
    text-align: center;
    margin-bottom: 10px;
}}
.report-title h1 {{
    font-size: 14pt;
    color: #1F497D;
    font-weight: bold;
}}
.report-title .period {{
    font-size: 9pt;
    color: #666;
    margin-top: 2px;
}}

/* ---- Section headings ---- */
h2 {{
    color: #366092;
    font-size: 11pt;
    margin: 12px 0 4px 0;
    padding-bottom: 2px;
    border-bottom: 2px solid #366092;
}}
h3 {{
    color: #4F81BD;
    font-size: 9pt;
    margin: 8px 0 2px 0;
}}

/* ---- Tables ---- */
table {{
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 6px;
    font-size: 8.5pt;
}}
th {{
    background: #366092;
    color: #fff;
    padding: 4px 6px;
    font-weight: bold;
    border: 1px solid #c0c0c0;
}}
.day-block table th {{
    background: #4F81BD;
}}
td {{
    padding: 3px 5px;
    border: 1px solid #e0e0e0;
}}
tr.even td {{ background: #F5F8FC; }}
tr.subtotal td {{
    background: #e2ecf4;
    font-weight: bold;
    border-top: 1.5px solid #366092;
    border-bottom: 1.5px solid #366092;
}}

/* Summary total row */
tfoot.grand-total td {{
    background: #D9E5F0;
    font-weight: bold;
    border-top: 2px solid #1F497D;
    border-bottom: 2px solid #1F497D;
}}

/* Alignment helpers */
.num  {{ text-align: right; white-space: nowrap; }}
.name {{ min-width: 100px; }}
.center {{ text-align: center; }}

</style>
</head>
<body>

<div class="first-page">
<div class="report-title">
    <h1>របាយការណ៍វត្តមាន និងប្រាក់ឈ្នួលបុគ្គលិក</h1>
    <p class="period">កាលបរិច្ឆេទ: {display_period}</p>
</div>

<h2>សរុបប្រាក់ឈ្នួល</h2>
<table>
    <thead>
        <tr>
            <th style="width:30px">ល.រ</th>
            <th>ឈ្មោះបុគ្គលិក</th>
            <th>ភេទ</th>
            <th class="num">តម្លៃថ្ងៃ</th>
            <th class="num">តម្លៃម៉ោង</th>
            <th class="num">ថ្ងៃធ្វើការ</th>
            <th class="num">ថែមម៉ោង</th>
            <th class="num">លុយសរុប</th>
            <th class="num">ខ្ចីសរុប</th>
            <th class="num">ប្រាក់កាត់</th>
            <th class="num">ត្រូវបើក ($)</th>
            <th class="num">ត្រូវបើក (៛)</th>
        </tr>
    </thead>
    <tbody>{summary_rows}</tbody>
    <tfoot class="grand-total">
        <tr>
            <td colspan="3">សរុប</td>
            <td></td>
            <td></td>
            <td class="num">{total_days_sum:.1f} ថ្ងៃ</td>
            <td class="num">{total_ot_sum:.1f} h</td>
            <td class="num">{int(round(total_gross_sum)):,} ៛</td>
            <td class="num">{int(round(total_borrow_sum)):,} ៛</td>
            <td class="num">{int(round(total_deduct_sum)):,} ៛</td>
            <td class="num">${total_net_usd_sum:.2f}</td>
            <td class="num">{int(round(total_net_khr_sum)):,} ៛</td>
        </tr>
    </tfoot>
</table>
</div>

<h2 style="page-break-before: always;">បំណែងចែកតាមថ្ងៃ</h2>
{daily_sections}

</body>
</html>"""


async def generate_pdf_report(reports_data: list, period_str: str, output_path: str, exchange_rate: float = 4000.0):
    """Generate a PDF report using Playwright (Chromium) for perfect Khmer text rendering.
    
    Playwright uses Chromium which has full HarfBuzz-based Khmer shaping.
    One-time setup: pip install playwright && playwright install chromium
    """
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        raise RuntimeError(
            "playwright is not installed.\n"
            "Run:\n"
            "  pip install playwright\n"
            "  playwright install chromium"
        )

    font_path = await ensure_khmer_font_async()
    html_content = _build_report_html(reports_data, period_str, font_path, exchange_rate)

    logger.info(f"Generating PDF with Playwright: {output_path}")
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-gpu"
            ]
        )
        page = await browser.new_page()
        await page.set_content(html_content, wait_until="domcontentloaded")
        await page.pdf(
            path=output_path,
            print_background=True,
            prefer_css_page_size=True,
        )
        await browser.close()
    logger.info("PDF report generated successfully.")





