import os
os.environ["DATABASE_URL"] = "sqlite:///test_attendance.db"
os.environ["SKIP_PDF_TEST"] = "1"
import unittest
from database import (
    init_db,
    add_employee,
    update_employee_name,
    delete_employee,
    get_employee_rate,
    get_all_employees,
    save_attendance_report,
    get_accumulated_totals,
    get_reports_by_dates,
    restart_attendance_count,
    Report,
    AttendanceRecord,
    SessionLocal,
    Base,
    record_borrow
)
from parser import parse_report_text_by_days
from report_generator import generate_excel_report, generate_pdf_report
from openpyxl import load_workbook

class TestIntegration(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        init_db()

    def setUp(self):
        # Clean up database tables before each test
        db = SessionLocal()
        try:
            db.execute(Base.metadata.tables['employees'].delete())
            db.execute(Base.metadata.tables['reports'].delete())
            db.execute(Base.metadata.tables['attendance'].delete())
            db.execute(Base.metadata.tables['settings'].delete())
            db.commit()
        finally:
            db.close()

    def test_employee_crud(self):
        # Add employee
        success = add_employee("ប៉ែន ទិត្យ", 80000)
        self.assertTrue(success)
        self.assertEqual(get_employee_rate("ប៉ែន ទិត្យ"), 80000)

        # Update employee rate
        success = add_employee("ប៉ែន ទិត្យ", 96000)
        self.assertTrue(success)
        self.assertEqual(get_employee_rate("ប៉ែន ទិត្យ"), 96000)

        # Rename employee
        success = update_employee_name("ប៉ែន ទិត្យ", "ប៉ែន ទិត្យថ្មី")
        self.assertTrue(success)
        self.assertIsNone(get_employee_rate("ប៉ែន ទិត្យ"))
        self.assertEqual(get_employee_rate("ប៉ែន ទិត្យថ្មី"), 96000)

        # Rename to already existing name (should fail)
        add_employee("អៀម អេន", 64000)
        success = update_employee_name("ប៉ែន ទិត្យថ្មី", "អៀម អេន")
        self.assertFalse(success)
        self.assertEqual(get_employee_rate("ប៉ែន ទិត្យថ្មី"), 96000)

        # List all
        all_emps = get_all_employees()
        self.assertEqual(len(all_emps), 2)
        self.assertEqual(all_emps["ប៉ែន ទិត្យថ្មី"]["rate"], 96000)
        self.assertEqual(all_emps["អៀម អេន"]["rate"], 64000)

        # Delete employee
        success = delete_employee("អៀម អេន")
        self.assertTrue(success)
        self.assertIsNone(get_employee_rate("អៀម អេន"))
        self.assertEqual(len(get_all_employees()), 1)

        # Test gender registration
        success = add_employee("ធិន", 30000, "ស")
        self.assertTrue(success)
        all_emps = get_all_employees()
        self.assertEqual(all_emps["ធិន"]["rate"], 30000)
        self.assertEqual(all_emps["ធិន"]["gender"], "ស")

    def test_bulk_add_parsing_logic(self):
        content = """
        ប៉ែន ទិត្យ 80000
        អៀម អេន 64000
        """
        lines = content.split("\n")
        parsed = []
        for line in lines:
            cleaned = line.strip()
            if not cleaned:
                continue
            parts = cleaned.split()
            rate = float(parts[-1])
            name = " ".join(parts[:-1]).strip()
            add_employee(name, rate)
            parsed.append((name, rate))
            
        self.assertEqual(len(parsed), 2)
        self.assertEqual(parsed[0], ("ប៉ែន ទិត្យ", 80000.0))
        self.assertEqual(parsed[1], ("អៀម អេន", 64000.0))
        
        all_emps = get_all_employees()
        self.assertEqual(all_emps["ប៉ែន ទិត្យ"]["rate"], 80000.0)
        self.assertEqual(all_emps["អៀម អេន"]["rate"], 64000.0)

    def test_report_processing(self):
        # Register workers with daily rates
        add_employee("ប៉ែន ទិត្យ", 80000) # 80000៛/day
        add_employee("អៀម អេន", 64000)   # 64000៛/day
        
        test_data = """
        1. ប៉ែន ទិត្យ.   8 h MEP
        2. អៀម អេន.     8.9 h
        3. ម៉ាច សិន 4 h
        """
        
        blocks = parse_report_text_by_days(test_data)
        self.assertEqual(len(blocks), 1)
        workers = blocks[0]['workers']
        self.assertEqual(len(workers), 3)
        
        results = {}
        for w in workers:
            name = w['name']
            hours = w['hours']
            
            daily_rate = get_employee_rate(name)
            rate = daily_rate if daily_rate is not None else 0.0
            
            salary = (hours / 8.0) * rate
            results[name] = {"hours": hours, "salary": salary}
            
        self.assertEqual(results["ប៉ែន ទិត្យ"]["hours"], 8.0)
        self.assertEqual(results["ប៉ែន ទិត្យ"]["salary"], 80000.0)
        self.assertEqual(results["អៀម អេន"]["hours"], 8.9)
        self.assertEqual(results["អៀម អេន"]["salary"], 71200.0)
        self.assertEqual(results["ម៉ាច សិន"]["hours"], 4.0)
        self.assertEqual(results["ម៉ាច សិន"]["salary"], 0.0)

    def test_multi_day_report_processing(self):
        add_employee("ប៉ែន ទិត្យ", 80000)
        add_employee("អៀម អេន", 64000)
        
        test_data = """
        Monday
        1. ប៉ែន ទិត្យ.   8 h MEP
        2. អៀម អេន.     8.9 h
        
        Tuesday
        1. ម៉ាច សិន 4 h
        2. ប៉ែន ទិត្យ.   8 h
        """
        
        blocks = parse_report_text_by_days(test_data)
        self.assertEqual(len(blocks), 2)
        
        worker_totals = {}
        for block in blocks:
            for w in block['workers']:
                name = w['name']
                hours = w['hours']
                
                daily_rate = get_employee_rate(name)
                rate = daily_rate if daily_rate is not None else 0.0
                
                salary = (hours / 8.0) * rate
                
                if name not in worker_totals:
                    worker_totals[name] = {'hours': 0.0, 'salary': 0.0}
                    
                worker_totals[name]['hours'] += hours
                worker_totals[name]['salary'] += salary
                
        self.assertEqual(worker_totals["ប៉ែន ទិត្យ"]["hours"], 16.0)
        self.assertEqual(worker_totals["ប៉ែន ទិត្យ"]["salary"], 160000.0)
        self.assertEqual(worker_totals["អៀម អេន"]["hours"], 8.9)
        self.assertEqual(worker_totals["អៀម អេន"]["salary"], 71200.0)
        self.assertEqual(worker_totals["ម៉ាច សិន"]["hours"], 4.0)
        self.assertEqual(worker_totals["ម៉ាច សិន"]["salary"], 0.0)

    def test_save_attendance_report_db(self):
        add_employee("ប៉ែន ទិត្យ", 80000)
        
        workers = [
            {'index': 1, 'name': "ប៉ែន ទិត្យ", 'hours': 8.0, 'note': "MEP"},
            {'index': 2, 'name': "ម៉ាច សិន", 'hours': 4.0, 'note': None}
        ]
        
        report_id = save_attendance_report("16-Jun-2026", "Monday", workers)
        self.assertIsNotNone(report_id)
        
        db = SessionLocal()
        try:
            report = db.query(Report).filter(Report.id == report_id).first()
            self.assertIsNotNone(report)
            self.assertEqual(report.date, "16-Jun-2026")
            self.assertEqual(report.header, "Monday")
            
            records = db.query(AttendanceRecord).filter(AttendanceRecord.report_id == report_id).all()
            self.assertEqual(len(records), 2)
            
            rec1 = next(r for r in records if r.employee_name == "ប៉ែន ទិត្យ")
            self.assertEqual(rec1.multiplier, 1.0)
            self.assertEqual(rec1.hours, 8.0)
            self.assertEqual(rec1.daily_rate, 80000.0)
            self.assertEqual(rec1.salary, 80000.0)
            self.assertEqual(rec1.note, "MEP")
            
            rec2 = next(r for r in records if r.employee_name == "ម៉ាច សិន")
            self.assertEqual(rec2.multiplier, 0.5)
            self.assertEqual(rec2.hours, 4.0)
            self.assertEqual(rec2.daily_rate, 0.0)
            self.assertEqual(rec2.salary, 0.0)
            self.assertIsNone(rec2.note)
        finally:
            db.close()

    def test_get_accumulated_totals(self):
        add_employee("ប៉ែន ទិត្យ", 80000)
        add_employee("អៀម អេន", 64000)
        
        workers1 = [
            {'index': 1, 'name': "ប៉ែន ទិត្យ", 'hours': 8.0, 'note': None},
            {'index': 2, 'name': "អៀម អេន", 'hours': 8.0, 'note': None}
        ]
        save_attendance_report("15-Jun-2026", "Monday", workers1)
        
        workers2 = [
            {'index': 1, 'name': "ប៉ែន ទិត្យ", 'hours': 4.0, 'note': None},
            {'index': 2, 'name': "អៀម អេន", 'hours': 10.0, 'note': None}
        ]
        save_attendance_report("16-Jun-2026", "Tuesday", workers2)
        
        totals = get_accumulated_totals()
        self.assertEqual(len(totals), 2)
        
        self.assertEqual(totals["ប៉ែន ទិត្យ"]["hours"], 12.0)
        self.assertEqual(totals["ប៉ែន ទិត្យ"]["salary"], 120000.0)
        self.assertEqual(totals["អៀម អេន"]["hours"], 18.0)
        self.assertEqual(totals["អៀម អេន"]["salary"], 144000.0)

    def test_report_overwrite_same_day(self):
        add_employee("ប៉ែន ទិត្យ", 80000)
        
        workers1 = [
            {'index': 1, 'name': "ប៉ែន ទិត្យ", 'hours': 8.0, 'note': None}
        ]
        id1 = save_attendance_report("16-Jun-2026", "ថ្ងៃទី: 16.06.26 (7:00am - 5:00pm)", workers1)
        
        workers2 = [
            {'index': 1, 'name': "ប៉ែន ទិត្យ", 'hours': 10.0, 'note': "Updated"}
        ]
        id2 = save_attendance_report("16-Jun-2026", "ថ្ងៃទី: 16.06.26 (8:00am - 6:00pm)", workers2)
        
        db = SessionLocal()
        try:
            reports = db.query(Report).all()
            self.assertEqual(len(reports), 1)
            self.assertEqual(reports[0].id, id2)
            self.assertEqual(reports[0].header, "ថ្ងៃទី: 16.06.26 (8:00am - 6:00pm)")
            
            records = db.query(AttendanceRecord).all()
            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].report_id, id2)
            self.assertEqual(records[0].hours, 10.0)
            self.assertEqual(records[0].note, "Updated")
        finally:
            db.close()

    def test_generate_excel_layout(self):
        # Add employee
        add_employee("ប៉ែន ទិត្យ", 80000)
        add_employee("អៀម អេន", 64000)
        
        # Save two days of reports
        workers_day1 = [
            {'index': 1, 'name': "ប៉ែន ទិត្យ", 'hours': 8.0, 'note': "MEP"},
            {'index': 2, 'name': "អៀម អេន", 'hours': 7.5, 'note': None}
        ]
        save_attendance_report("01-Jun-2026", "ថ្ងៃទី: 1.06.26 (7:00am - 5:00pm)", workers_day1)
        
        workers_day2 = [
            {'index': 1, 'name': "ប៉ែន ទិត្យ", 'hours': 6.0, 'note': "Half day"},
        ]
        save_attendance_report("02-Jun-2026", "ថ្ងៃទី: 2.06.26", workers_day2)
        
        # Retrieve reports
        reports_data = get_reports_by_dates("1.06.26", "2.06.26")
        self.assertEqual(len(reports_data), 2)
        
        output_path = "tmp/test_report.xlsx"
        os.makedirs("tmp", exist_ok=True)
        if os.path.exists(output_path):
            os.remove(output_path)
            
        generate_excel_report(reports_data, output_path, 4000.0)
        
        self.assertTrue(os.path.exists(output_path))
        
        # Load workbook and check sheet contents
        wb = load_workbook(output_path)
        self.assertIn("Summary", wb.sheetnames)
        self.assertIn("Details", wb.sheetnames)
        
        ws = wb["Details"]
        
        # Check title
        self.assertEqual(ws["A1"].value, "ព័ត៌មានលម្អិតប្រចាំថ្ងៃ")
        
        # Check header row (row 3)
        self.assertEqual(ws.cell(row=3, column=1).value, "លរ")
        self.assertEqual(ws.cell(row=3, column=2).value, "ឈ្មោះ")
        self.assertEqual(ws.cell(row=3, column=3).value, "ភេទ")
        self.assertEqual(ws.cell(row=3, column=4).value, "តម្លៃ(៛)/ថ្ងៃ")
        self.assertEqual(ws.cell(row=3, column=5).value, "តម្លៃ(៛)/ម៉ោង")
        
        self.assertEqual(ws.cell(row=3, column=6).value, "1") # Day 1 Date
        self.assertEqual(ws.cell(row=3, column=10).value, "2") # Day 2 Date
        
        # Check subheaders row (row 4)
        self.assertEqual(ws.cell(row=4, column=6).value, "Day")
        self.assertEqual(ws.cell(row=4, column=7).value, "OT(h)")
        self.assertEqual(ws.cell(row=4, column=8).value, "Borrow")
        self.assertEqual(ws.cell(row=4, column=9).value, "Project")
        
        self.assertEqual(ws.cell(row=3, column=14).value, "សរុបថ្ងៃធ្វើការ")
        self.assertEqual(ws.cell(row=3, column=15).value, "សរុបថែមម៉ោង")
        self.assertEqual(ws.cell(row=3, column=16).value, "លុយសរុប")
        self.assertEqual(ws.cell(row=3, column=17).value, "លុយបានខ្ចី សរុប")
        self.assertEqual(ws.cell(row=3, column=18).value, "ប្រាក់កាត់")
        self.assertEqual(ws.cell(row=3, column=19).value, "លុយត្រូវបើក")
        
        # Check employee names
        emp1_name = ws.cell(row=5, column=2).value
        emp2_name = ws.cell(row=6, column=2).value
        
        names = [emp1_name, emp2_name]
        self.assertIn("ប៉ែន ទិត្យ", names)
        self.assertIn("អៀម អេន", names)
        
        # Verify font and style for names and values
        self.assertEqual(ws.cell(row=5, column=2).font.name, "Times New Roman")
        self.assertEqual(ws.cell(row=6, column=2).font.name, "Times New Roman")
        
        pen_row = 5 if emp1_name == "ប៉ែន ទិត្យ" else 6
        iam_row = 6 if pen_row == 5 else 5
        
        # Check values
        self.assertEqual(ws.cell(row=pen_row, column=4).value, 80000.0) # Rate/day
        self.assertEqual(ws.cell(row=pen_row, column=5).value, f"=D{pen_row}/8") # Rate/hour
        self.assertEqual(ws.cell(row=pen_row, column=6).value, 1.0) # Day 1 base day
        self.assertEqual(ws.cell(row=pen_row, column=7).value, 0.0) # Day 1 OT
        self.assertEqual(ws.cell(row=pen_row, column=10).value, 0.75) # Day 2 base day
        self.assertEqual(ws.cell(row=pen_row, column=11).value, 0.0) # Day 2 OT
        
        self.assertEqual(ws.cell(row=pen_row, column=14).value, f"=F{pen_row}+J{pen_row}") # Total days worked formula
        self.assertEqual(ws.cell(row=pen_row, column=15).value, f"=G{pen_row}+K{pen_row}") # Total OT formula
        self.assertEqual(ws.cell(row=pen_row, column=16).value, f"=(N{pen_row}*D{pen_row})+(O{pen_row}*E{pen_row})") # Total salary formula
        self.assertEqual(ws.cell(row=pen_row, column=17).value, f"=H{pen_row}+L{pen_row}") # Borrow formula
        self.assertEqual(ws.cell(row=pen_row, column=19).value, f"=P{pen_row}-Q{pen_row}-R{pen_row}") # Net formula
        
        self.assertEqual(ws.cell(row=iam_row, column=4).value, 64000.0)
        self.assertEqual(ws.cell(row=iam_row, column=6).value, 7.5 / 8.0)
        self.assertEqual(ws.cell(row=iam_row, column=7).value, 0.0)
        self.assertIn(ws.cell(row=iam_row, column=10).value, [None, ""]) # Did not work day 2
        
        # Total Row (row 7)
        self.assertEqual(ws.cell(row=7, column=1).value, "សរុប")
        self.assertEqual(ws.cell(row=7, column=6).value, "=SUM(F5:F6)")
        self.assertEqual(ws.cell(row=7, column=7).value, "=SUM(G5:G6)")
        self.assertEqual(ws.cell(row=7, column=14).value, "=SUM(N5:N6)")
        self.assertEqual(ws.cell(row=7, column=16).value, "=SUM(P5:P6)")
        self.assertEqual(ws.cell(row=7, column=19).value, "=SUM(S5:S6)")
 
        # Verify summary sheet properties
        ws_sum = wb["Summary"]
        self.assertEqual(ws_sum["A1"].font.name, "Times New Roman")
        self.assertEqual(ws_sum["A2"].font.name, "Times New Roman")
        self.assertEqual(ws_sum.cell(row=4, column=1).font.name, "Times New Roman")
        self.assertEqual(ws_sum.cell(row=5, column=1).font.name, "Times New Roman")
        
        # Cleanup
        if os.path.exists(output_path):
            os.remove(output_path)

    def test_generate_pdf_layout(self):
        if os.getenv("SKIP_PDF_TEST"):
            raise unittest.SkipTest("Skipping PDF test in sandbox environment")
        import asyncio
        add_employee("ប៉ែន ទិត្យ", 80000)
        
        workers = [{'index': 1, 'name': "ប៉ែន ទិត្យ", 'hours': 8.0, 'note': "MEP"}]
        save_attendance_report("15-Jun-2026", "ថ្ងៃទី: 15.06.26", workers)
        
        reports_data = get_reports_by_dates("15.06.26", "15.06.26")
        output_path = "tmp/test_report.pdf"
        os.makedirs("tmp", exist_ok=True)
        if os.path.exists(output_path):
            os.remove(output_path)
            
        asyncio.run(generate_pdf_report(reports_data, "15.06.26", output_path, 4000.0))
        self.assertTrue(os.path.exists(output_path))
        if os.path.exists(output_path):
            os.remove(output_path)

    def test_exchange_rate_crud(self):
        from database import get_exchange_rate, set_exchange_rate
        # Default rate should be 4000
        self.assertEqual(get_exchange_rate(), 4000.0)

        # Set new rate
        set_exchange_rate(4100.0)
        self.assertEqual(get_exchange_rate(), 4100.0)

        # Reset rate
        set_exchange_rate(4000.0)

    def test_employee_cascade_sync(self):
        # 1. Add employee
        add_employee("ប៉ែន ទិត្យ", 80000)
        
        # 2. Add attendance record for employee
        workers = [
            {'index': 1, 'name': "ប៉ែន ទិត្យ", 'hours': 8.0, 'note': "Test"}
        ]
        save_attendance_report("16-Jun-2026", "Monday", workers)
        
        # Verify it has KHR rate and 80000 salary
        db = SessionLocal()
        try:
            rec = db.query(AttendanceRecord).filter(AttendanceRecord.employee_name == "ប៉ែន ទិត្យ").first()
            self.assertEqual(rec.daily_rate, 80000.0)
            self.assertEqual(rec.salary, 80000.0)
        finally:
            db.close()
            
        # 3. Update daily rate of employee
        add_employee("ប៉ែន ទិត្យ", 96000)
        
        # Verify attendance record updated
        db = SessionLocal()
        try:
            rec = db.query(AttendanceRecord).filter(AttendanceRecord.employee_name == "ប៉ែន ទិត្យ").first()
            self.assertEqual(rec.daily_rate, 96000.0)
            self.assertEqual(rec.salary, 96000.0)
        finally:
            db.close()

        # 4. Rename employee
        update_employee_name("ប៉ែន ទិត្យ", "ប៉ែន ទិត្យថ្មី")
        
        # Verify attendance record name is renamed
        db = SessionLocal()
        try:
            rec_old = db.query(AttendanceRecord).filter(AttendanceRecord.employee_name == "ប៉ែន ទិត្យ").first()
            rec_new = db.query(AttendanceRecord).filter(AttendanceRecord.employee_name == "ប៉ែន ទិត្យថ្មី").first()
            self.assertIsNone(rec_old)
            self.assertIsNotNone(rec_new)
            self.assertEqual(rec_new.daily_rate, 96000.0)
        finally:
            db.close()

        # 5. Delete employee
        delete_employee("ប៉ែន ទិត្យថ្មី")
        
        # Verify attendance record is deleted
        db = SessionLocal()
        try:
            rec = db.query(AttendanceRecord).filter(AttendanceRecord.employee_name == "ប៉ែន ទិត្យថ្មី").first()
            self.assertIsNone(rec)
        finally:
            db.close()

    def test_bulk_delete_logic(self):
        add_employee("ប៉ែន ទិត្យ", 80000)
        add_employee("អៀម អេន", 64000)
        
        # Verify they exist
        self.assertEqual(len(get_all_employees()), 2)
        
        # Bulk delete
        names_to_delete = ["ប៉ែន ទិត្យ", "អៀម អេន", "ម៉ាច សិន"]
        deleted = []
        for name in names_to_delete:
            if delete_employee(name):
                deleted.append(name)
                
        self.assertEqual(len(deleted), 2)
        self.assertIn("ប៉ែន ទិត្យ", deleted)
        self.assertIn("អៀម អេន", deleted)
        self.assertEqual(len(get_all_employees()), 0)

    def test_restart_attendance_count(self):
        # 1. Add employee (this should NOT be deleted)
        add_employee("ប៉ែន ទិត្យ", 80000)
        
        # 2. Add attendance reports (these SHOULD be deleted)
        workers = [
            {'index': 1, 'name': "ប៉ែន ទិត្យ", 'hours': 8.0, 'note': "Test"}
        ]
        save_attendance_report("16-Jun-2026", "Monday", workers)
        
        # Verify db has records and reports
        db = SessionLocal()
        try:
            self.assertEqual(db.query(Report).count(), 1)
            self.assertEqual(db.query(AttendanceRecord).count(), 1)
            self.assertEqual(len(get_all_employees()), 1)
        finally:
            db.close()
            
        # 3. Call restart_attendance_count
        success = restart_attendance_count()
        self.assertTrue(success)
        
        # Verify reports and attendance are empty, but employee remains
        db = SessionLocal()
        try:
            self.assertEqual(db.query(Report).count(), 0)
            self.assertEqual(db.query(AttendanceRecord).count(), 0)
            self.assertEqual(len(get_all_employees()), 1)
        finally:
            db.close()

    def test_borrow_feature(self):
        # 1. Setup Employee
        add_employee("ប៉ែន ទិត្យ", 80000)
        
        # 2. Save Attendance Report
        from datetime import timezone, timedelta, datetime
        tz_kh = timezone(timedelta(hours=7))
        today_str = datetime.now(tz_kh).strftime("%d-%b-%Y")
        
        workers = [
            {'index': 1, 'name': "ប៉ែន ទិត្យ", 'hours': 8.0, 'note': "MEP"}
        ]
        save_attendance_report(today_str, f"ថ្ងៃទី: {today_str} (07:00 AM - 05:00 PM)", workers)
        
        # 3. Test record_borrow with deduction thresholds
        # Threshold 1: <= 200,000 => deduction = 0
        success, name, r_day = record_borrow("ប៉ែន ទិត្យ", 150000, 0)
        self.assertTrue(success)
        self.assertEqual(name, "ប៉ែន ទិត្យ")
        
        # Verify database record note
        db = SessionLocal()
        try:
            rec = db.query(AttendanceRecord).filter(AttendanceRecord.employee_name == "ប៉ែន ទិត្យ").first()
            self.assertIn("ខ្ចី 150000", rec.note)
            self.assertNotIn("កាត់", rec.note)
            self.assertIn("MEP", rec.note) # original note preserved
        finally:
            db.close()
            
        # Threshold 2: > 200,000 and <= 400,000 => deduction = 20,000
        success, name, r_day = record_borrow("ប៉ែន ទិត្យ", 250000, 20000)
        self.assertTrue(success)
        db = SessionLocal()
        try:
            rec = db.query(AttendanceRecord).filter(AttendanceRecord.employee_name == "ប៉ែន ទិត្យ").first()
            self.assertIn("ខ្ចី 250000", rec.note)
            self.assertIn("កាត់ 20000", rec.note)
            self.assertNotIn("150000", rec.note) # old borrow replaced
        finally:
            db.close()

        # Threshold 3: > 400,000 => deduction = 40,000
        success, name, r_day = record_borrow("ប៉ែន ទិត្យ", 450000, 40000)
        self.assertTrue(success)
        db = SessionLocal()
        try:
            rec = db.query(AttendanceRecord).filter(AttendanceRecord.employee_name == "ប៉ែន ទិត្យ").first()
            self.assertIn("ខ្ចី 450000", rec.note)
            self.assertIn("កាត់ 40000", rec.note)
            self.assertNotIn("250000", rec.note)
            self.assertNotIn("20000", rec.note)
        finally:
            db.close()

        # Test carrying over on re-submission of report
        # Re-submit report without borrow note in the input workers list
        workers_re = [
            {'index': 1, 'name': "ប៉ែន ទិត្យ", 'hours': 9.0, 'note': "MEP"}
        ]
        save_attendance_report(today_str, f"ថ្ងៃទី: {today_str} (07:00 AM - 05:00 PM)", workers_re)
        
        db = SessionLocal()
        try:
            rec = db.query(AttendanceRecord).filter(AttendanceRecord.employee_name == "ប៉ែន ទិត្យ").first()
            self.assertEqual(rec.hours, 9.0)
            self.assertIn("ខ្ចី 450000", rec.note)
            self.assertIn("កាត់ 40000", rec.note)
            self.assertIn("MEP", rec.note)
        finally:
            db.close()
            
        # Test clearing borrow (amount = 0, deduction = 0)
        success, name, r_day = record_borrow("ប៉ែន ទិត្យ", 0, 0)
        self.assertTrue(success)
        db = SessionLocal()
        try:
            rec = db.query(AttendanceRecord).filter(AttendanceRecord.employee_name == "ប៉ែន ទិត្យ").first()
            self.assertEqual(rec.note, "MEP") # cleanly reverted to original note
        finally:
            db.close()

if __name__ == '__main__':
    unittest.main()
